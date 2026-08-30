"""
generate_descriptions.py -- rewrite the "Description (optional)" column with
short, plain-language shop copy generated from the inventory's own scent
family / top notes data.

Not pulled from Fragrantica or anywhere else -- same reasoning as the
storefront's generated bottle-glyph work (see app/components/bottle-glyph.tsx):
their write-ups are copyrighted editorial content, not ours to republish.

One short sentence per product on purpose: the customer base is a passing
mall kiosk shopper, not a fragrance forum -- an original version of this
script wrote a full note-pyramid breakdown ("Opens with rosewood, cardamom
and sichuan pepper, moves through oud and sandalwood, settles into tonka
bean and vanilla...") plus a second sentence of gender/concentration
marketing copy that just duplicated what the product page's own spec line
and selectors already show. That read as connoisseur copy, not shelf copy,
and buried the one plain, evocative line (the family phrase) a quick-glance
shopper actually wants. This keeps just that line plus, when there's a top
note to name, one recognizable highlight -- not all 7 notes across three
tiers.

Usage:
    python generate_descriptions.py --file Perfume_Inventory_100.xlsx --dry-run
    python generate_descriptions.py --file Perfume_Inventory_100.xlsx --write-xlsx
    python generate_descriptions.py --file Perfume_Inventory_100.xlsx --sql-out update_descriptions.sql
"""

import argparse
import hashlib

import pandas as pd

from sku import make_sku

SHEET_NAME = "Inventory"
HEADER_ROW = 1

# A handful of varied phrasings per scent family so products in the same
# family don't all read identically -- rotated deterministically by SKU
# (see `pick`), not randomly, so re-runs are stable.
FAMILY_PHRASES = {
    "Citrus": [
        "Bright and zesty from the first spray, effervescent rather than sweet",
        "Sharp and sparkling, built for warm afternoons",
        "Sunlit and clean, never heavy",
    ],
    "Fresh/Aquatic": [
        "Cool and marine, like sea spray on skin",
        "Clean and breezy, the fragrance equivalent of an open window",
        "Airy and transparent, built to feel weightless",
    ],
    "Green": [
        "Crisp and leafy, like something just cut from the garden",
        "Verdant and cool, with a bitter-green backbone",
        "Grassy and sharp, unmistakably outdoors",
    ],
    "Aromatic": [
        "Herbal and crisp, with a peppery, energetic edge",
        "Sharp and green-tinged, built for movement",
        "Brisk and confident, the classic aromatic snap",
    ],
    "Fougère": [
        "Classically barbershop-fresh, lavender over ferny green",
        "Structured and cool, a modern take on an old formula",
        "Crisp and herbal with a soapy, clean-shaven finish",
    ],
    "Floral": [
        "Soft-petaled and radiant, built around a true floral heart",
        "Delicate but never quiet, a bouquet that lingers",
        "Romantic and airy, blooming outward as it wears",
    ],
    "Chypre": [
        "Mossy and complex, old-world in the best way",
        "Earthy and refined, with real depth to it",
        "Structured and sophisticated, rewards a second sniff",
    ],
    "Woody": [
        "Warm and grounded, built around a real wood backbone",
        "Smooth and confident, fills a room quietly",
        "Dry and textured, more sweater than cologne",
    ],
    "Amber/Oriental": [
        "Warm and resinous, glowing rather than shouting",
        "Rich and enveloping, built for cooler nights",
        "Spiced and sensual, the kind of scent people lean in for",
    ],
    "Gourmand": [
        "Sweet and edible without tipping into dessert",
        "Warm and comforting, like something baked",
        "Indulgent and cozy, built around real sweetness",
    ],
    "Leather": [
        "Smoky and supple, real leather rather than an idea of it",
        "Dry and worn-in, with real edge to it",
        "Dark and textured, with some grit to it",
    ],
    "Spicy": [
        "Warm-spiced and alive, with real heat to it",
        "Peppery and confident, built to be noticed",
        "Richly spiced, the opposite of quiet",
    ],
}

DEFAULT_FAMILY_PHRASE = "Distinctive and well-built, a scent with real character"


def pick(pool: dict, key: str, sku: str):
    """Deterministic rotation through a phrase pool, keyed by SKU so the
    same product gets the same phrasing every time this is re-run."""
    options = pool.get(key)
    if not options:
        return None
    index = int(hashlib.sha256(sku.encode()).hexdigest(), 16) % len(options)
    return options[index]


def parse_notes(value) -> list[str]:
    if pd.isna(value) or not str(value).strip():
        return []
    return [n.strip() for n in str(value).split(",") if n.strip()]


def build_description(row, sku: str) -> str:
    family = str(row.get("Scent Family") or "").strip()
    family_phrase = pick(FAMILY_PHRASES, family, sku) or DEFAULT_FAMILY_PHRASE

    # Just the first listed top note, not all three top notes plus heart
    # and base -- one recognizable ingredient is enough to make a product
    # feel distinct from its shelf neighbors without turning the blurb into
    # a note-pyramid breakdown. Spreadsheet order is treated as the
    # product's own "lead" note, same assumption the old pyramid sentence
    # made by listing top notes first.
    top_notes = parse_notes(row.get("Top Notes"))
    highlight = top_notes[0].strip().lower() if top_notes else None

    if highlight:
        return f"{family_phrase}, with a hint of {highlight}."
    return f"{family_phrase}."


def load(file: str) -> pd.DataFrame:
    df = pd.read_excel(file, sheet_name=SHEET_NAME, header=HEADER_ROW)
    df.columns = [c.strip() for c in df.columns]
    df = df[df["Brand"].notna() & df["Product Name"].notna() & df["Size"].notna()].copy()
    # make_sku() from etl/sku.py -- see that file for why this replaced a
    # third independent copy of the same slug-and-join logic.
    df["SKU"] = df.apply(
        lambda r: make_sku(r["Brand"], r["Product Name"], r.get("Concentration"), r["Size"]),
        axis=1,
    )
    return df.reset_index(drop=True)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True)
    ap.add_argument("--dry-run", action="store_true", help="preview only, no writes")
    ap.add_argument("--write-xlsx", action="store_true", help="overwrite the Description column in the workbook")
    ap.add_argument("--sql-out", help="write a SQL file that UPDATEs dim_products.description by sku")
    args = ap.parse_args()

    df = load(args.file)
    df["New Description"] = df.apply(lambda r: build_description(r, r["SKU"]), axis=1)

    print(f"{len(df)} products.\n")
    sample = df.sample(min(6, len(df)), random_state=1)
    for _, r in sample.iterrows():
        print(f"{r['Brand']} {r['Product Name']} ({r['SKU']})")
        print(f"  old: {r['Description (optional)']}")
        print(f"  new: {r['New Description']}\n")

    if args.dry_run:
        return

    if args.sql_out:
        with open(args.sql_out, "w", encoding="utf-8") as f:
            f.write("-- Generated by generate_descriptions.py -- shop copy built from the\n")
            f.write("-- inventory's own note/family/gender/concentration data, not scraped.\n")
            f.write("BEGIN;\n")
            for _, r in df.iterrows():
                escaped = r["New Description"].replace("'", "''")
                f.write(
                    f"UPDATE dim_products SET description = '{escaped}', updated_at = now() "
                    f"WHERE sku = '{r['SKU']}';\n"
                )
            f.write("COMMIT;\n")
        print(f"Wrote {args.sql_out} ({len(df)} statements).")

    if args.write_xlsx:
        import openpyxl

        wb = openpyxl.load_workbook(args.file)
        ws = wb[SHEET_NAME]
        header_row_idx = HEADER_ROW + 1  # openpyxl rows are 1-indexed
        headers = {cell.value: cell.column for cell in ws[header_row_idx]}
        desc_col = headers["Description (optional)"]
        brand_col, name_col, size_col = headers["Brand"], headers["Product Name"], headers["Size"]
        concentration_col = headers.get("Concentration")

        new_by_sku = dict(zip(df["SKU"], df["New Description"]))
        updated = 0
        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            brand = ws.cell(row_idx, brand_col).value
            name = ws.cell(row_idx, name_col).value
            size = ws.cell(row_idx, size_col).value
            concentration = ws.cell(row_idx, concentration_col).value if concentration_col else None
            if not brand or not name or not size:
                continue
            # make_sku(), not the old bare slug(f"{brand}-{name}-{size}") --
            # that local slug() was removed when this file switched to the
            # shared etl/sku.py (see load() above), which left this branch
            # calling an undefined name. Same function, same fields dry-run
            # and --write-xlsx now use, or the SKUs used to key new_by_sku
            # here silently stop matching the ones load() produced.
            sku = make_sku(brand, name, concentration, size)
            if sku in new_by_sku:
                ws.cell(row_idx, desc_col).value = new_by_sku[sku]
                updated += 1
        wb.save(args.file)
        print(f"Updated {updated} descriptions in {args.file}.")


if __name__ == "__main__":
    main()
